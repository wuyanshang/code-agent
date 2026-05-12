import csv
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest

from code_agent.tools.table_file_tools import ReadTableFileTool, read_csv_file


class MockContext:
    def __init__(self, project_root: Path) -> None:
        self.project_root = project_root

        class Config:
            class Tools:
                read_file_max_chars = 50000

            tools = Tools()

        self.config = Config()

    class PathGuard:
        def __init__(self, root: Path) -> None:
            self.root = root

        def ensure_allowed(self, path: str) -> Path:
            return self.root / path

    @property
    def path_guard(self):
        return self.PathGuard(self.project_root)


def test_read_csv_simple():
    """测试读取简单CSV文件"""
    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        csv_file = project_root / "test.csv"

        # 创建CSV文件
        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Name", "Age", "City"])
            writer.writerow(["Alice", "25", "Beijing"])
            writer.writerow(["Bob", "30", "Shanghai"])

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "test.csv"}, context)

        assert result.ok
        assert "Name" in result.content
        assert "Alice" in result.content
        assert "Bob" in result.content
        assert "Beijing" in result.content


def test_read_csv_with_max_rows():
    """测试限制读取行数"""
    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        csv_file = project_root / "large.csv"

        # 创建大CSV文件
        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "Value"])
            for i in range(200):
                writer.writerow([str(i), f"value_{i}"])

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "large.csv", "max_rows": 10}, context)

        assert result.ok
        assert "显示前10行" in result.content
        # 应该只包含前几行
        assert "value_0" in result.content
        assert "value_1" in result.content
        # 不应该包含后面的行
        assert "value_50" not in result.content


def test_read_csv_empty():
    """测试读取空CSV文件"""
    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        csv_file = project_root / "empty.csv"
        csv_file.write_text("", encoding="utf-8")

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "empty.csv"}, context)

        assert result.ok
        assert "空" in result.content


def test_read_excel_simple():
    """测试读取Excel文件"""
    pytest.importorskip("openpyxl")

    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        excel_file = project_root / "test.xlsx"

        # 创建Excel文件
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Name", "Age", "City"])
        ws.append(["Alice", 25, "Beijing"])
        ws.append(["Bob", 30, "Shanghai"])
        wb.save(excel_file)

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "test.xlsx"}, context)

        assert result.ok
        assert "Name" in result.content
        assert "Alice" in result.content
        assert "Bob" in result.content


def test_read_excel_multiple_sheets():
    """测试读取多工作表Excel"""
    pytest.importorskip("openpyxl")

    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        excel_file = project_root / "multi.xlsx"

        # 创建多工作表Excel
        from openpyxl import Workbook

        wb = Workbook()

        # 第一个工作表
        ws1 = wb.active
        ws1.title = "Users"
        ws1.append(["Name", "Age"])
        ws1.append(["Alice", 25])

        # 第二个工作表
        ws2 = wb.create_sheet("Products")
        ws2.append(["Product", "Price"])
        ws2.append(["Apple", 5])

        wb.save(excel_file)

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "multi.xlsx"}, context)

        assert result.ok
        assert "Users" in result.content
        assert "Products" in result.content
        assert "Alice" in result.content
        assert "Apple" in result.content


def test_read_unsupported_file():
    """测试读取不支持的文件类型"""
    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        txt_file = project_root / "test.txt"
        txt_file.write_text("hello", encoding="utf-8")

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "test.txt"}, context)

        assert not result.ok
        assert "不支持" in result.error


def test_read_nonexistent_file():
    """测试读取不存在的文件"""
    with TemporaryDirectory() as tmpdir:
        context = MockContext(Path(tmpdir))
        tool = ReadTableFileTool()

        result = tool.execute({"path": "nonexistent.csv"}, context)

        assert not result.ok
        assert "不存在" in result.error


def test_csv_with_chinese():
    """测试包含中文的CSV"""
    with TemporaryDirectory() as tmpdir:
        project_root = Path(tmpdir)
        csv_file = project_root / "chinese.csv"

        # 创建包含中文的CSV
        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["姓名", "年龄", "城市"])
            writer.writerow(["张三", "25", "北京"])
            writer.writerow(["李四", "30", "上海"])

        context = MockContext(project_root)
        tool = ReadTableFileTool()

        result = tool.execute({"path": "chinese.csv"}, context)

        assert result.ok
        assert "姓名" in result.content
        assert "张三" in result.content
        assert "北京" in result.content
