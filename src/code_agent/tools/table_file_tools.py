from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from code_agent.schemas import ToolResult
from code_agent.tools.base import BaseTool


def read_csv_file(file_path: Path, max_rows: int = 100) -> str:
    """
    读取CSV文件并格式化为表格。

    Args:
        file_path: CSV文件路径
        max_rows: 最大读取行数

    Returns:
        格式化的表格字符串
    """
    try:
        with open(file_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)

        if not rows:
            return "CSV文件为空"

        # 计算每列的最大宽度
        col_widths = []
        if rows:
            num_cols = max(len(row) for row in rows)
            for col_idx in range(num_cols):
                max_width = 0
                for row in rows:
                    if col_idx < len(row):
                        max_width = max(max_width, len(str(row[col_idx])))
                col_widths.append(min(max_width, 50))  # 限制最大宽度

        # 格式化输出
        lines = []
        lines.append(f"CSV文件: {file_path.name}")
        lines.append(f"总行数: {len(rows)}" + (f" (显示前{max_rows}行)" if len(rows) >= max_rows else ""))
        lines.append("")

        # 表头（如果第一行看起来像表头）
        if rows:
            header = rows[0]
            header_line = " | ".join(
                str(cell).ljust(col_widths[i]) if i < len(col_widths) else str(cell)
                for i, cell in enumerate(header)
            )
            lines.append(header_line)
            lines.append("-" * len(header_line))

            # 数据行
            for row in rows[1:]:
                row_line = " | ".join(
                    str(cell).ljust(col_widths[i]) if i < len(col_widths) else str(cell)
                    for i, cell in enumerate(row)
                )
                lines.append(row_line)

        return "\n".join(lines)

    except UnicodeDecodeError:
        # 尝试其他编码
        try:
            with open(file_path, "r", encoding="gbk", newline="") as f:
                reader = csv.reader(f)
                rows = list(reader)[:max_rows]
            return f"CSV文件（GBK编码）: {file_path.name}\n总行数: {len(rows)}\n\n" + "\n".join(
                ", ".join(row) for row in rows
            )
        except Exception as e:
            return f"无法读取CSV文件（编码错误）: {e}"
    except Exception as e:
        return f"读取CSV文件失败: {e}"


def read_excel_file(file_path: Path, max_rows: int = 100) -> str:
    """
    读取Excel文件并格式化为表格。

    Args:
        file_path: Excel文件路径
        max_rows: 最大读取行数

    Returns:
        格式化的表格字符串
    """
    try:
        import openpyxl
    except ImportError:
        return "错误: 需要安装openpyxl库才能读取Excel文件\n运行: pip install openpyxl"

    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        lines = []
        lines.append(f"Excel文件: {file_path.name}")
        lines.append(f"工作表数量: {len(workbook.sheetnames)}")
        lines.append("")

        # 读取所有工作表
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            lines.append(f"=== 工作表: {sheet_name} ===")

            # 读取数据
            rows = []
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                # 过滤掉全空行
                if any(cell is not None for cell in row):
                    rows.append(row)

            if not rows:
                lines.append("(空工作表)")
                lines.append("")
                continue

            lines.append(f"行数: {len(rows)}" + (f" (显示前{max_rows}行)" if len(rows) >= max_rows else ""))
            lines.append("")

            # 计算列宽
            col_widths = []
            if rows:
                num_cols = max(len(row) for row in rows)
                for col_idx in range(num_cols):
                    max_width = 0
                    for row in rows:
                        if col_idx < len(row) and row[col_idx] is not None:
                            max_width = max(max_width, len(str(row[col_idx])))
                    col_widths.append(min(max_width, 50))

            # 格式化输出
            for row_idx, row in enumerate(rows):
                row_line = " | ".join(
                    str(cell if cell is not None else "").ljust(col_widths[i])
                    if i < len(col_widths)
                    else str(cell if cell is not None else "")
                    for i, cell in enumerate(row)
                )
                lines.append(row_line)

                # 表头分隔线
                if row_idx == 0:
                    lines.append("-" * len(row_line))

            lines.append("")

        workbook.close()
        return "\n".join(lines)

    except Exception as e:
        return f"读取Excel文件失败: {e}"


class ReadTableFileTool(BaseTool):
    name = "read_table_file"
    description = (
        "读取表格文件（CSV、Excel）并以表格形式显示内容。"
        "支持.csv、.xlsx、.xls格式。"
    )
    parameters_schema = {
        "type": "object",
        "properties": {
            "path": {"type": "string", "description": "文件路径"},
            "max_rows": {
                "type": "integer",
                "description": "最大读取行数，默认100",
                "default": 100,
            },
        },
        "required": ["path"],
    }

    def execute(self, arguments: dict[str, Any], context: Any) -> ToolResult:
        path = context.path_guard.ensure_allowed(arguments["path"])

        if not path.exists():
            return ToolResult(ok=False, content="", error=f"文件不存在: {path}")

        if not path.is_file():
            return ToolResult(ok=False, content="", error=f"不是文件: {path}")

        max_rows = int(arguments.get("max_rows", 100))
        suffix = path.suffix.lower()

        # 根据文件类型读取
        if suffix == ".csv":
            content = read_csv_file(path, max_rows)
        elif suffix in (".xlsx", ".xls"):
            content = read_excel_file(path, max_rows)
        else:
            return ToolResult(
                ok=False,
                content="",
                error=f"不支持的文件类型: {suffix}\n支持的类型: .csv, .xlsx, .xls",
            )

        # 限制输出长度
        max_chars = context.config.tools.read_file_max_chars
        if len(content) > max_chars:
            content = content[:max_chars] + f"\n\n...[输出过长，已截断，共{len(content)}字符]"

        return ToolResult(ok=True, content=content)
